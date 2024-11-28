import pandas as pd
import os
from datetime import datetime
from openpyxl import load_workbook

# Caminho do arquivo TXT original
path_txt = r"\\WKRadar-SRV\Relatórios\TI-PBI_EstoqueLote.txt"
current_dir = os.getcwd()

# Base para os arquivos numerados
base_name = "1.xlsx"

# Remover arquivos numerados antigos e criar o novo arquivo
for i in range(1, 6):
    file_path = os.path.join(current_dir, f"{i}.xlsx")
    if os.path.exists(file_path):
        os.remove(file_path)
output_excel = os.path.join(current_dir, base_name)

# Tentativa de leitura do arquivo TXT com delimitador correto
try:
    df = pd.read_csv(path_txt, sep='|', encoding='ISO-8859-1')
except UnicodeDecodeError:
    print("Erro ao ler o arquivo. Verifique a codificação ou o formato do arquivo.")
    exit()

# Renomeando as colunas relevantes
df = df.rename(columns={
    "Código Produto": "Código do produto",
    "Nome do Produto": "Nome",
    "Marca": "Marca",
    "Inativo": "Inativo"
})

# Selecionando apenas as colunas necessárias
df = df[["Código do produto", "Nome", "Marca", "Inativo"]]

# Somando os produtos repetidos
df_grouped = df.groupby(["Código do produto", "Nome", "Marca", "Inativo"], as_index=False).size()

# Renomeando a coluna com a soma
df_grouped = df_grouped.rename(columns={"size": "Quantidade disponivel"})

# Salvando o resultado em um arquivo Excel
df_grouped.to_excel(output_excel, index=False)

# Ajustando a formatação do Excel
try:
    wb = load_workbook(output_excel)
    ws = wb.active

    # Aplicar formatação personalizada para a coluna "Vezes de repetição"
    for cell in ws["E"]:  # Supondo que "Vezes de repetição" esteja na coluna E
        if isinstance(cell.value, (int, float)):  # Certifica-se de que é um número
            cell.number_format = '0"x"'  # Formato personalizado, como 1x, 2x

    # Salvar o arquivo formatado
    wb.save(output_excel)
    print(f"Planilha '{output_excel}' formatada com sucesso na pasta do código.")
except Exception as e:
    print(f"Erro ao aplicar a formatação no Excel: {e}")
