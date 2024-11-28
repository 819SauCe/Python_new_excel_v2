# Como ultimo passo estilizar a planilha para ficar 100% entendivel
# Design moderno, e simples
# ajustar conforme preferir

import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment


def estilizar_planilha(caminho_arquivo):
    try:
        workbook = load_workbook(caminho_arquivo)
        sheet = workbook.active

        # Definição de estilos
        cor_fundo = PatternFill(start_color='feffba', end_color='feffba', fill_type='solid')
        cor_vinho = PatternFill(start_color='509e4a', end_color='509e4a', fill_type='solid')
        fonte_branca = Font(color='000000')
        fonte_branca_super = Font(color='000000', bold=True)

        fundo_verde = PatternFill(start_color='32a852', end_color='32a852', fill_type='solid')
        fonte_verde = Font(color='013d11', bold=True)
        fundo_amarelo = PatternFill(start_color='bec23e', end_color='bec23e', fill_type='solid')
        fonte_amarelo = Font(color='7e8200', bold=True)
        fundo_vermelho = PatternFill(start_color='ba0b0b', end_color='ba0b0b', fill_type='solid')
        fonte_vermelho = Font(color='610000', bold=True)

        borda_branca = Border(left=Side(style='thin', color='a6a6a6'),
                              right=Side(style='thin', color='a6a6a6'),
                              top=Side(style='thin', color='a6a6a6'),
                              bottom=Side(style='thin', color='a6a6a6'))

        # Identificar as colunas pelo cabeçalho
        colunas = {cell.value: cell.column for cell in sheet[1]}  # Mapear cabeçalhos
        coluna_inativo = colunas.get('Inativo', None)

        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
            for cell in row:
                cell.fill = cor_fundo
                cell.font = fonte_branca
                cell.border = borda_branca

                if cell.column == 1:
                    cell.alignment = Alignment(horizontal='center')

                # Estilização baseada em valores A, B e C
                if cell.value == 'A':
                    cell.fill = fundo_verde
                    cell.font = fonte_verde
                elif cell.value == 'B':
                    cell.fill = fundo_amarelo
                    cell.font = fonte_amarelo
                elif cell.value == 'C':
                    cell.fill = fundo_vermelho
                    cell.font = fonte_vermelho

                # Estilização da coluna "Inativo"
                if coluna_inativo and cell.column == coluna_inativo:
                    if str(cell.value).strip().lower() == 'ativo':
                        cell.fill = fundo_verde
                        cell.font = fonte_verde
                    elif str(cell.value).strip().lower() == 'inativo':
                        cell.fill = fundo_vermelho
                        cell.font = fonte_vermelho

                # Estilização para "Sim" e "Não"
                if str(cell.value).strip().lower() == 'sim':
                    cell.fill = fundo_verde
                    cell.font = fonte_verde
                elif str(cell.value).strip().lower() == 'não':
                    cell.fill = fundo_vermelho
                    cell.font = fonte_vermelho

        # Cabeçalhos da planilha
        for cell in sheet[1]:
            cell.fill = cor_vinho
            cell.font = fonte_branca_super

        # Ajuste automático da largura das colunas
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value:
                    text_length = len(str(cell.value))
                    column_letter = cell.column_letter
                    current_width = sheet.column_dimensions[column_letter].width or 0
                    estimated_width = (text_length + 2)
                    sheet.column_dimensions[column_letter].width = max(current_width, estimated_width)

        # Aplicar filtro na tabela
        sheet.auto_filter.ref = sheet.dimensions

        # Fechar e salvar o workbook
        workbook.save(caminho_arquivo)
        workbook.close()  # Garante que o arquivo seja liberado

        print(f"Estilização aplicada com sucesso ao arquivo: {caminho_arquivo}")

    except PermissionError:
        print(f"Erro: o arquivo '{caminho_arquivo}' está em uso. Feche-o e tente novamente.")
    except Exception as e:
        print(f"Erro ao estilizar a planilha: {e}")


def aplicar_estilizacao_nas_planilhas():
    # Diretório raiz (onde o script está sendo executado)
    current_dir = os.getcwd()

    # Listar arquivos esperados (1.xlsx, 2.xlsx, etc.)
    arquivos_excel = [file for file in os.listdir(current_dir) if file.endswith('.xlsx') and file in ["1.xlsx", "2.xlsx"]]

    if not arquivos_excel:
        print("Nenhum arquivo compatível encontrado para estilização.")
        return

    # Aplicar estilização no arquivo mais recente (ordem crescente de nomes)
    for arquivo in sorted(arquivos_excel):
        caminho_arquivo = os.path.join(current_dir, arquivo)
        print(f"Aplicando estilização no arquivo: {caminho_arquivo}")
        estilizar_planilha(caminho_arquivo)


if __name__ == "__main__":
    aplicar_estilizacao_nas_planilhas()
